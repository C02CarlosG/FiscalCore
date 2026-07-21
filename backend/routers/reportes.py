from __future__ import annotations

import re
from decimal import Decimal
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from .. import db, isr, iva
from ..deps import get_current_user, validar_acceso_empresa, serializar

router = APIRouter(tags=["Reportes"])


def _excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _header_row(ws, values: list) -> None:
    ws.append(values)
    bold = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = bold


def _dec(value) -> Optional[float]:
    return float(value) if value is not None else None


@router.get("/api/v1/empresas/{empresa_id}/reportes/conciliacion/{periodo}")
async def reporte_conciliacion(
    empresa_id: str,
    periodo: str,
    current_user: dict = Depends(get_current_user),
):
    validar_acceso_empresa(empresa_id, current_user)

    rows = db.query_all(
        """
        SELECT con.tipo_match, con.monto_movimiento, con.monto_cfdi,
               con.diferencia, con.porcentaje_match, con.notas, con.confianza,
               cf.uuid      AS cfdi_uuid,
               cf.rfc_emisor,
               cf.fecha_emision
        FROM conciliaciones con
        LEFT JOIN cfdi cf ON cf.id = con.cfdi_id
        WHERE con.empresa_id = %s AND con.periodo = %s
        ORDER BY con.tipo_match, con.monto_movimiento DESC NULLS LAST
        """,
        (empresa_id, periodo),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliación"
    _header_row(ws, [
        "Tipo Match", "Monto Movimiento", "Monto CFDI", "Diferencia",
        "% Match", "Confianza", "Notas", "UUID CFDI", "RFC Emisor", "Fecha Emisión",
    ])
    for r in rows:
        ws.append([
            r.get("tipo_match", ""),
            _dec(r.get("monto_movimiento")),
            _dec(r.get("monto_cfdi")),
            _dec(r.get("diferencia")),
            _dec(r.get("porcentaje_match")),
            r.get("confianza", ""),
            r.get("notas", ""),
            r.get("cfdi_uuid", ""),
            r.get("rfc_emisor", ""),
            str(r["fecha_emision"])[:10] if r.get("fecha_emision") else "",
        ])

    return _excel_response(wb, f"conciliacion_{periodo}.xlsx")


@router.get("/api/v1/empresas/{empresa_id}/reportes/riesgos/{periodo}")
async def reporte_riesgos(
    empresa_id: str,
    periodo: str,
    current_user: dict = Depends(get_current_user),
):
    validar_acceso_empresa(empresa_id, current_user)

    rows = db.query_all(
        """
        SELECT r.codigo      AS tipo_riesgo,
               r.nombre,
               r.severidad,
               d.descripcion,
               d.monto_afectado,
               d.estado,
               d.created_at AS fecha_deteccion
        FROM detecciones d
        JOIN riesgos r ON r.id = d.riesgo_id
        WHERE d.empresa_id = %s AND d.periodo = %s
        ORDER BY
            CASE r.severidad
                WHEN 'critico' THEN 1
                WHEN 'alto'    THEN 2
                WHEN 'medio'   THEN 3
                ELSE 4
            END,
            d.created_at DESC
        """,
        (empresa_id, periodo),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riesgos"
    _header_row(ws, [
        "Tipo Riesgo", "Nombre", "Severidad", "Descripción",
        "Monto Afectado", "Estado", "Fecha Detección",
    ])
    for r in rows:
        ws.append([
            r.get("tipo_riesgo", ""),
            r.get("nombre", ""),
            r.get("severidad", ""),
            r.get("descripcion", ""),
            _dec(r.get("monto_afectado")),
            r.get("estado", ""),
            str(r["fecha_deteccion"])[:10] if r.get("fecha_deteccion") else "",
        ])

    return _excel_response(wb, f"riesgos_{periodo}.xlsx")


@router.get("/api/v1/empresas/{empresa_id}/reportes/scoring/{periodo}")
async def reporte_scoring(
    empresa_id: str,
    periodo: str,
    current_user: dict = Depends(get_current_user),
):
    validar_acceso_empresa(empresa_id, current_user)

    row = db.query_one(
        "SELECT * FROM scoring_fiscal WHERE empresa_id = %s AND periodo = %s",
        (empresa_id, periodo),
    )
    if not row:
        raise HTTPException(status_code=404, detail="Sin scoring para este período")
    return serializar(row)


@router.get("/api/v1/empresas/{empresa_id}/diot/{periodo}")
async def generar_diot(
    empresa_id: str,
    periodo: str,
    current_user: dict = Depends(get_current_user),
):
    validar_acceso_empresa(empresa_id, current_user)

    rows = db.query_all(
        """
        SELECT c.rfc_emisor   AS rfc_proveedor,
               c.nombre_emisor AS nombre,
               SUM(c.subtotal)       AS monto_total,
               SUM(c.iva_trasladado) AS iva_pagado,
               COUNT(*)              AS num_facturas
        FROM cfdi c
        JOIN empresas e ON e.id = c.empresa_id
        WHERE c.empresa_id = %s
          AND c.rfc_receptor = e.rfc
          AND c.fecha_emision >= (%s || '-01')::date
          AND c.fecha_emision  < ((%s || '-01')::date + INTERVAL '1 month')
          AND c.estado = 'vigente'
        GROUP BY c.rfc_emisor, c.nombre_emisor
        ORDER BY monto_total DESC
        """,
        (empresa_id, periodo, periodo),
    )

    return {
        "periodo":           periodo,
        "total_proveedores": len(rows),
        "registros": [
            {
                "rfc_proveedor":  r["rfc_proveedor"],
                "nombre":         r["nombre"] or "",
                "tipo_operacion": "03",
                "monto_total":    float(r["monto_total"] or 0),
                "iva_pagado":     float(r["iva_pagado"] or 0),
                "num_facturas":   r["num_facturas"],
            }
            for r in rows
        ],
    }


# ---------------------------------------------------------------------------
# Cédula de IVA (Módulo 3) — cálculo por flujo de efectivo (Art. 1-B LIVA)
# ---------------------------------------------------------------------------

_PERIODO_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


def _floats(obj):
    """Convierte recursivamente Decimal -> float en dicts/listas anidados."""
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, dict):
        return {k: _floats(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_floats(v) for v in obj]
    return obj


def _cargar_datos_cedula_iva(empresa_id: str, periodo: str):
    """Carga de DB los insumos de la cédula: RFC, CFDIs candidatos, pagos del
    periodo y el IVA acreditable devengado (base DIOT) para el comparativo."""
    emp = db.query_one("SELECT rfc FROM empresas WHERE id = %s", (empresa_id,))
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    rfc = emp["rfc"]

    cfdis = db.query_all(
        """
        SELECT uuid, tipo_comprobante, metodo_pago, estado, es_anticipo_sat,
               rfc_emisor, rfc_receptor, forma_pago, fecha_emision,
               subtotal, total, iva_trasladado
        FROM cfdi
        WHERE empresa_id = %s
          AND estado = 'vigente'
          AND (
                metodo_pago = 'PPD'
                OR (fecha_emision >= (%s || '-01')::date
                    AND fecha_emision  < ((%s || '-01')::date + INTERVAL '1 month'))
              )
        """,
        (empresa_id, periodo, periodo),
    )

    pagos = db.query_all(
        """
        SELECT pr.cfdi_uuid, pr.importe_pagado, p.fecha_pago
        FROM pagos_cfdi p
        JOIN pagos_relaciones pr ON pr.pago_id = p.id
        WHERE p.empresa_id = %s
          AND p.fecha_pago >= (%s || '-01')::date
          AND p.fecha_pago  < ((%s || '-01')::date + INTERVAL '1 month')
        """,
        (empresa_id, periodo, periodo),
    )

    diot = db.query_one(
        """
        SELECT COALESCE(SUM(c.iva_trasladado), 0) AS iva
        FROM cfdi c
        JOIN empresas e ON e.id = c.empresa_id
        WHERE c.empresa_id = %s
          AND c.rfc_receptor = e.rfc
          AND c.estado = 'vigente'
          AND c.fecha_emision >= (%s || '-01')::date
          AND c.fecha_emision  < ((%s || '-01')::date + INTERVAL '1 month')
        """,
        (empresa_id, periodo, periodo),
    )
    diot_iva = (diot["iva"] if diot and diot["iva"] is not None else Decimal("0"))
    return rfc, cfdis, pagos, Decimal(str(diot_iva))


@router.get("/api/v1/empresas/{empresa_id}/cedula-iva/{periodo}")
async def cedula_iva(
    empresa_id: str,
    periodo: str,
    factor: float = 1.0,   # factor de prorrateo (Art. 5-V); 1.0 = 100% gravado
    current_user: dict = Depends(get_current_user),
):
    """Cédula mensual de IVA por flujo de efectivo: trasladado, acreditable,
    prorrateo, resultado del periodo y comparativo contra el IVA devengado (DIOT)."""
    if not _PERIODO_RE.match(periodo):
        raise HTTPException(status_code=422, detail="periodo inválido; formato esperado YYYY-MM")
    validar_acceso_empresa(empresa_id, current_user)

    rfc, cfdis, pagos, diot_iva = _cargar_datos_cedula_iva(empresa_id, periodo)

    trasladado = iva.iva_trasladado(cfdis, pagos, periodo, rfc)
    acred = iva.iva_acreditable(cfdis, pagos, periodo, rfc)
    factor_dec = Decimal(str(factor))
    ajustado = iva.aplicar_prorrateo(acred["bruto"], factor_dec)

    iva_retenido = Decimal("0.00")  # v1: retenciones no computadas todavía
    por_pagar = (trasladado["total"] - ajustado - iva_retenido).quantize(Decimal("0.01"))

    resultado = {
        "iva_por_pagar": por_pagar,
        "saldo_a_cargo": por_pagar if por_pagar > 0 else Decimal("0.00"),
        "saldo_a_favor": -por_pagar if por_pagar < 0 else Decimal("0.00"),
    }

    return _floats({
        "empresa_id": empresa_id,
        "periodo": periodo,
        "trasladado": trasladado,
        "acreditable": {**acred, "factor_prorrateo": factor_dec, "ajustado": ajustado},
        "iva_retenido": iva_retenido,
        "resultado": resultado,
        "comparativo_sat": {
            "diot_iva_pagado": diot_iva,
            "diferencia": (ajustado - diot_iva).quantize(Decimal("0.01")),
        },
    })


# ---------------------------------------------------------------------------
# ISR provisional (Módulo 5) — pago provisional mensual, base devengado (Art. 14 LISR)
# ---------------------------------------------------------------------------


def _cargar_datos_isr(empresa_id: str, periodo: str):
    """Carga de DB los insumos del pago provisional: config anual (CU/PTU/pérdidas/
    tasa), el ingreso nominal acumulado del ejercicio a cada corte mensual (ene..mes)
    y la retención de ISR de CADA mes del ejercicio hasta el declarado (no solo el
    mes declarado: ``isr.isr_provisional`` recalcula el pago real de los meses
    anteriores, y ese pago real ya viene reducido por su propia retención — ver
    fix de Kilo Code Review en PR #4). Devuelve ``(None, {}, {})`` si no hay
    ``config_isr_empresa`` para el ejercicio (falta el CU)."""
    emp = db.query_one("SELECT rfc FROM empresas WHERE id = %s", (empresa_id,))
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")

    ejercicio = periodo[:4]
    mes = int(periodo[5:7])

    config = db.query_one(
        """
        SELECT coeficiente_utilidad, tasa_isr, ptu_pagada, perdidas_pendientes
        FROM config_isr_empresa
        WHERE empresa_id = %s AND ejercicio = %s
        """,
        (empresa_id, ejercicio),
    )
    if not config:
        return None, {}, {}

    filas = db.query_all(
        """
        SELECT EXTRACT(MONTH FROM c.fecha_emision)::int AS mes,
               COALESCE(SUM(
                 CASE WHEN c.tipo_comprobante = 'E' THEN -c.subtotal ELSE c.subtotal END
               ), 0) AS ingreso
        FROM cfdi c
        JOIN empresas e ON e.id = c.empresa_id
        WHERE c.empresa_id = %s
          AND c.rfc_emisor = e.rfc
          AND c.tipo_comprobante IN ('I','E')
          AND c.estado = 'vigente'
          AND c.es_anticipo_sat = FALSE
          AND c.fecha_emision >= (%s || '-01-01')::date
          AND c.fecha_emision  < ((%s || '-01')::date + INTERVAL '1 month')
        GROUP BY EXTRACT(MONTH FROM c.fecha_emision)
        """,
        (empresa_id, ejercicio, periodo),
    )
    ingreso_del_mes = {int(f["mes"]): Decimal(str(f["ingreso"])) for f in filas}
    ingresos_por_mes: dict[int, Decimal] = {}
    acumulado = Decimal("0")
    for m in range(1, mes + 1):
        acumulado += ingreso_del_mes.get(m, Decimal("0"))
        ingresos_por_mes[m] = acumulado

    filas_retencion = db.query_all(
        """
        SELECT EXTRACT(MONTH FROM c.fecha_emision)::int AS mes,
               COALESCE(SUM(c.isr_retenido), 0) AS isr_retenido
        FROM cfdi c
        JOIN empresas e ON e.id = c.empresa_id
        WHERE c.empresa_id = %s
          AND c.rfc_emisor = e.rfc
          AND c.estado = 'vigente'
          AND c.fecha_emision >= (%s || '-01-01')::date
          AND c.fecha_emision  < ((%s || '-01')::date + INTERVAL '1 month')
        GROUP BY EXTRACT(MONTH FROM c.fecha_emision)
        """,
        (empresa_id, ejercicio, periodo),
    )
    retenciones_por_mes = {
        int(f["mes"]): Decimal(str(f["isr_retenido"])) for f in filas_retencion
    }

    return config, ingresos_por_mes, retenciones_por_mes


@router.get("/api/v1/empresas/{empresa_id}/isr-provisional/{periodo}")
async def isr_provisional_endpoint(
    empresa_id: str,
    periodo: str,
    current_user: dict = Depends(get_current_user),
):
    """Pago provisional mensual de ISR (persona moral, régimen general, Art. 14 LISR):
    acumulado del ejercicio a la fecha de corte, base devengado."""
    if not _PERIODO_RE.match(periodo):
        raise HTTPException(status_code=422, detail="periodo inválido; formato esperado YYYY-MM")
    validar_acceso_empresa(empresa_id, current_user)

    config, ingresos_por_mes, retenciones_por_mes = _cargar_datos_isr(empresa_id, periodo)
    if not config:
        raise HTTPException(
            status_code=404,
            detail="Sin coeficiente de utilidad configurado para este ejercicio",
        )

    mes = int(periodo[5:7])
    cu = Decimal(str(config["coeficiente_utilidad"]))
    tasa = Decimal(str(config["tasa_isr"]))
    ptu = Decimal(str(config["ptu_pagada"]))
    perdidas = Decimal(str(config["perdidas_pendientes"]))

    calculo = isr.isr_provisional(ingresos_por_mes, mes, cu, tasa, ptu, perdidas, retenciones_por_mes)

    return _floats({
        "empresa_id": empresa_id,
        "periodo": periodo,
        "ejercicio": int(periodo[:4]),
        "coeficiente_utilidad": cu,
        "ingreso_nominal_acumulado": calculo["ingreso_nominal_acum"],
        "utilidad_estimada": calculo["utilidad_estimada"],
        "deducciones_base": {"ptu_pagada": ptu, "perdidas_pendientes": perdidas},
        "base_gravable": calculo["base_gravable"],
        "tasa_isr": tasa,
        "isr_acumulado": calculo["isr_acumulado"],
        "pagos_provisionales_anteriores": calculo["pagos_previos"],
        "isr_retenido": calculo["isr_retenido"],
        "resultado": {"pago_del_mes": calculo["pago_del_mes"]},
    })
