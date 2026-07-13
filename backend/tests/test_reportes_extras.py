"""Tests de los endpoints de reportes.py sin cédula propia: exportes a Excel
(conciliación, riesgos), scoring y DIOT (Día 22).

DB mockeada directamente sobre `backend.db` (estos endpoints no tienen un
loader dedicado como los de las cédulas, hacen `db.query_all`/`db.query_one`
en línea).
"""
from decimal import Decimal

import openpyxl
from fastapi.testclient import TestClient
from io import BytesIO

import backend.main_api as main
from backend import db
from backend.deps import get_current_user
from backend.routers import reportes

client = TestClient(main.app)


def _auth(monkeypatch):
    main.app.dependency_overrides[get_current_user] = lambda: {"id": "u1", "user_id": "u1"}
    monkeypatch.setattr(reportes, "validar_acceso_empresa", lambda *a, **k: None)


# ─── GET reportes/conciliacion/{periodo} (Excel) ──────────────────────────────

def test_reporte_conciliacion_genera_excel_con_filas(monkeypatch):
    _auth(monkeypatch)
    monkeypatch.setattr(db, "query_all", lambda *a, **k: [{
        "tipo_match": "exacto",
        "monto_movimiento": Decimal("1000.00"),
        "monto_cfdi": Decimal("1000.00"),
        "diferencia": Decimal("0.00"),
        "porcentaje_match": Decimal("100.0"),
        "notas": "match perfecto",
        "confianza": "alta",
        "cfdi_uuid": "UUID-1",
        "rfc_emisor": "PROV010101AAA",
        "fecha_emision": "2026-01-10",
    }])

    try:
        r = client.get("/api/v1/empresas/emp-1/reportes/conciliacion/2026-01")
    finally:
        main.app.dependency_overrides.clear()

    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.title == "Conciliación"
    assert [c.value for c in ws[1]][:3] == ["Tipo Match", "Monto Movimiento", "Monto CFDI"]
    fila = [c.value for c in ws[2]]
    assert fila[0] == "exacto"
    assert fila[1] == 1000.0
    assert fila[7] == "UUID-1"


def test_reporte_conciliacion_sin_filas_solo_encabezado(monkeypatch):
    _auth(monkeypatch)
    monkeypatch.setattr(db, "query_all", lambda *a, **k: [])

    try:
        r = client.get("/api/v1/empresas/emp-1/reportes/conciliacion/2026-01")
    finally:
        main.app.dependency_overrides.clear()

    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.max_row == 1


# ─── GET reportes/riesgos/{periodo} (Excel) ───────────────────────────────────

def test_reporte_riesgos_genera_excel_con_filas(monkeypatch):
    _auth(monkeypatch)
    monkeypatch.setattr(db, "query_all", lambda *a, **k: [{
        "tipo_riesgo": "gastos_sin_cfdi",
        "nombre": "Gasto sin CFDI",
        "severidad": "alto",
        "descripcion": "Cargo bancario sin CFDI relacionado",
        "monto_afectado": Decimal("500.00"),
        "estado": "abierto",
        "fecha_deteccion": "2026-01-15",
    }])

    try:
        r = client.get("/api/v1/empresas/emp-1/reportes/riesgos/2026-01")
    finally:
        main.app.dependency_overrides.clear()

    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.title == "Riesgos"
    fila = [c.value for c in ws[2]]
    assert fila[0] == "gastos_sin_cfdi"
    assert fila[2] == "alto"
    assert fila[4] == 500.0


# ─── GET reportes/scoring/{periodo} ───────────────────────────────────────────

def test_reporte_scoring_encontrado(monkeypatch):
    _auth(monkeypatch)
    monkeypatch.setattr(db, "query_one", lambda *a, **k: {
        "empresa_id": "emp-1", "periodo": "2026-01", "score_total": Decimal("85.5"),
    })

    try:
        r = client.get("/api/v1/empresas/emp-1/reportes/scoring/2026-01")
    finally:
        main.app.dependency_overrides.clear()

    assert r.status_code == 200
    assert r.json()["score_total"] == 85.5


def test_reporte_scoring_sin_datos_da_404(monkeypatch):
    _auth(monkeypatch)
    monkeypatch.setattr(db, "query_one", lambda *a, **k: None)

    try:
        r = client.get("/api/v1/empresas/emp-1/reportes/scoring/2026-01")
    finally:
        main.app.dependency_overrides.clear()

    assert r.status_code == 404


# ─── GET diot/{periodo} ────────────────────────────────────────────────────────

def test_generar_diot_agrupa_por_proveedor(monkeypatch):
    _auth(monkeypatch)
    monkeypatch.setattr(db, "query_all", lambda *a, **k: [{
        "rfc_proveedor": "PROV010101AAA",
        "nombre": "Proveedor Uno",
        "monto_total": Decimal("10000.00"),
        "iva_pagado": Decimal("1600.00"),
        "num_facturas": 3,
    }])

    try:
        r = client.get("/api/v1/empresas/emp-1/diot/2026-01")
    finally:
        main.app.dependency_overrides.clear()

    assert r.status_code == 200
    body = r.json()
    assert body["total_proveedores"] == 1
    reg = body["registros"][0]
    assert reg["rfc_proveedor"] == "PROV010101AAA"
    assert reg["monto_total"] == 10000.0
    assert reg["iva_pagado"] == 1600.0
    assert reg["num_facturas"] == 3
    assert reg["tipo_operacion"] == "03"


def test_generar_diot_sin_facturas_retorna_lista_vacia(monkeypatch):
    _auth(monkeypatch)
    monkeypatch.setattr(db, "query_all", lambda *a, **k: [])

    try:
        r = client.get("/api/v1/empresas/emp-1/diot/2026-01")
    finally:
        main.app.dependency_overrides.clear()

    assert r.status_code == 200
    body = r.json()
    assert body["total_proveedores"] == 0
    assert body["registros"] == []
